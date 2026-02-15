from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from src.constants import ColumnNames, ColumnPatterns
from src.header_detector import HeaderDetector


class HoldingsFileHandler:
    @classmethod
    def read(cls, input_file_path: Path | str) -> pd.DataFrame:
        print(f"Reading: {input_file_path}")
        raw_df = pd.read_excel(input_file_path, header=None)
        header_row = HeaderDetector.find_header_row(raw_df)
        holdings_df = pd.read_excel(input_file_path, header=header_row)
        holdings_df = cls._standardize_columns(holdings_df)
        holdings_df = cls._clean_input_data(holdings_df)
        print(f"Found {len(holdings_df)} holdings")
        return holdings_df

    @classmethod
    def _standardize_columns(cls, df: pd.DataFrame) -> pd.DataFrame:
        column_map = HeaderDetector.map_columns(df.columns)
        selected_df = df[
            [
                column_map[ColumnPatterns.CURRENCY],
                column_map[ColumnPatterns.PERCENT],
                column_map[ColumnPatterns.COUNTRY],
                column_map[ColumnPatterns.SECTOR],
            ]
        ].copy()
        selected_df.columns = [
            ColumnNames.CURRENCY,
            ColumnNames.PERCENT,
            ColumnNames.COUNTRY,
            ColumnNames.SECTOR,
        ]
        return selected_df

    @staticmethod
    def _clean_input_data(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df[ColumnNames.PERCENT] = pd.to_numeric(
            df[ColumnNames.PERCENT], errors="coerce"
        )
        df = df[df[ColumnNames.PERCENT] > 0]
        text_columns = [ColumnNames.CURRENCY, ColumnNames.COUNTRY, ColumnNames.SECTOR]
        for col in text_columns:
            df[col] = df[col].fillna("Unknown").astype(str).str.strip()
        return df

    @classmethod
    def write(cls, output_path: Path, df: pd.DataFrame) -> None:
        print("Writing output...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Aggregated Holdings"
        cls._write_formatted_data_to_worksheet(ws, df)
        cls._adjust_column_widths(ws)
        wb.save(output_path)
        print(f"Output saved to: {output_path}")

    @classmethod
    def _write_formatted_data_to_worksheet(
        cls, ws: Worksheet, df: pd.DataFrame
    ) -> None:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cls._format_header_cell(cell)
                    continue
                if c_idx == 5:
                    cls._format_weight_cell(cell)

    @staticmethod
    def _format_header_cell(cell: Cell | MergedCell) -> None:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _format_weight_cell(cell: Cell | MergedCell) -> None:
        cell.number_format = "0.00000%"
        if isinstance(cell.value, (int, float)):
            cell.value = cell.value / 100

    @staticmethod
    def _adjust_column_widths(ws: Worksheet):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
